from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def generate_xlsx(version, template_path):
    wb = load_workbook(template_path)
    ws = wb.active

    # Preencher cabeçalho
    ws['B2'] = version.budget.codigo_legado
    ws['D2'] = version.criado_em.strftime("%d/%m/%Y")

    row = 5  # Linha inicial para itens

    for budget_item in version.budgetitem_set.select_related('item'):
        item = budget_item.item
        ws[f'B{row}'] = item.artigo
        ws[f'C{row}'] = item.descricao
        ws[f'D{row}'] = budget_item.quantidade
        ws[f'E{row}'] = item.get_preco_final(**budget_item.atributos_personalizados)

        # Estilizar se houver atributos faltantes
        if not item.atributos_validos(budget_item.atributos_personalizados):
            ws.row_dimensions[row].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        row += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()