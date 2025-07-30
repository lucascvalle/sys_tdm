from django.http import HttpResponse
from django.conf import settings
import openpyxl
import io
from copy import copy
from openpyxl.utils import get_column_letter
from django.shortcuts import redirect
from django.contrib import messages

def exportar_consumo_material_excel(request, consumos_agregados, filtros):
    template_path = settings.BASE_DIR / 'static' / 'excel_templates' / 'modelo_consumo_material.xlsx'

    try:
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        # Preencher cabeçalho do relatório com filtros aplicados
        # Assumindo que os rótulos estão na coluna A do template
        sheet['B4'] = filtros.get('ref_obra', 'N/A')
        sheet['B5'] = filtros.get('data_inicio_ficha', 'N/A')
        sheet['B6'] = filtros.get('previsao_entrega_ficha', 'N/A')

        # Configurar cabeçalhos da tabela (linha 8)
        sheet.merge_cells('A8:C8')
        sheet['A8'] = "Materiais/Componentes"
        sheet['D8'] = "QTD"
        sheet['E8'] = "Tipo Un"

        # Inserir dados na tabela a partir da linha 9, saltando uma linha
        current_row = 9
        for consumo in consumos_agregados:
            sheet.merge_cells(f'A{current_row}:C{current_row}')
            componente_display = f"{consumo['componente__nome']}"
            if consumo['descricao_detalhada']:
                componente_display += f" - {consumo['descricao_detalhada']}"
            sheet.cell(row=current_row, column=1, value=componente_display)
            sheet.cell(row=current_row, column=4, value=float(consumo['total_quantidade']))
            sheet.cell(row=current_row, column=5, value=consumo['unidade'])
            current_row += 2  # Salta uma linha

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_consumo_material.xlsx"'
        return response

    except FileNotFoundError:
        messages.error(request, "O arquivo de template Excel (modelo_consumo_material.xlsx) não foi encontrado. Certifique-se de que está em sys_tdm/static/excel_templates/.")
        return redirect('consumos:material_consumption_report')
    except Exception as e:
        messages.error(request, f"Erro ao exportar relatório para Excel: {e}")
        return redirect('consumos:material_consumption_report')

def exportar_utilizacao_maquina_excel(request, sessoes_trabalho, filtros):
    template_path = settings.BASE_DIR / 'static' / 'excel_templates' / 'modelo_ficha_postos_maquinas.xlsx'

    try:
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        # Preencher cabeçalho do relatório
        sheet['B4'] = filtros.get('posto_trabalho', 'N/A')
        sheet['B5'] = filtros.get('data', 'N/A')

        # Configurar cabeçalhos da tabela (linha 7)
        # Assumindo que os rótulos já estão no template na linha 7

        # Inserir dados na tabela a partir da linha 8
        current_row = 8
        for sessao in sessoes_trabalho:
            sheet.cell(row=current_row, column=1, value=sessao.operador.nome)
            sheet.cell(row=current_row, column=2, value=sessao.ficha_obra.ref_obra if sessao.ficha_obra else "N/A")
            sheet.cell(row=current_row, column=3, value=sessao.operacao)
            sheet.cell(row=current_row, column=4, value=sessao.hora_inicio.strftime('%H:%M'))
            sheet.cell(row=current_row, column=5, value=sessao.hora_saida.strftime('%H:%M') if sessao.hora_saida else '--')
            current_row += 1

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="relatorio_utilizacao_maquina.xlsx"'
        return response

    except FileNotFoundError:
        messages.error(request, "O arquivo de template Excel (modelo_ficha_postos_maquinas.xlsx) não foi encontrado. Certifique-se de que está em sys_tdm/static/excel_templates/.")
        return redirect('consumos:machine_utilization_report')
    except Exception as e:
        messages.error(request, f"Erro ao exportar relatório para Excel: {e}")
        return redirect('consumos:machine_utilization_report')
