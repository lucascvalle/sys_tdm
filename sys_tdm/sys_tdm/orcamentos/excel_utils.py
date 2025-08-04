"""Utility functions for exporting budget and production data to Excel files.

This module contains the core logic for:
- Copying cell values and styles in OpenPyXL.
- Formatting detailed item descriptions based on product instances and configurations.
- Sanitizing names for use in template variables.
- Rendering instance and configuration descriptions using Django templates.
- Exporting budget data to Excel, including hierarchical grouping.
- Exporting production sheet data to Excel.
"""

from __future__ import annotations
import io
import re
import decimal
from copy import copy
from collections import defaultdict
from typing import Any, Dict, List, TYPE_CHECKING

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
from django.template import Template, Context

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils.translation import gettext_lazy as _

# Importações de modelos necessárias
from orcamentos.models import Orcamento, ItemOrcamento
from produtos.models import (
    ProdutoInstancia, ProdutoTemplate, Categoria, Atributo, ProdutoConfiguracao,
    InstanciaAtributo, InstanciaComponente, Componente, TemplateComponente,
    ConfiguracaoComponenteEscolha
)

# Type checking for HttpRequest to avoid circular imports if needed
if TYPE_CHECKING:
    from django.http import HttpRequest


def copy_cell(source_cell: openpyxl.cell.cell.Cell, target_cell: openpyxl.cell.cell.Cell) -> None:
    """
    Copia valor e estilo de forma defensiva de uma célula de origem para uma célula de destino.

    Garante que `number_format` nunca seja `None` para evitar erros.

    Args:
        source_cell: A célula de onde copiar o valor e o estilo.
        target_cell: A célula para onde copiar o valor e o estilo.
    """
    target_cell.value = source_cell.value if source_cell.value is not None else ""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format or 'General'


def copy_style(source_cell: openpyxl.cell.cell.Cell, target_cell: openpyxl.cell.cell.Cell) -> None:
    """
    Copia apenas o estilo de forma defensiva de uma célula de origem para uma célula de destino.

    Garante que `number_format` nunca seja `None`.

    Args:
        source_cell: A célula de onde copiar o estilo.
        target_cell: A célula para onde copiar o estilo.
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format or 'General'


def _format_detailed_item_description_base(item: ItemOrcamento, include_monetary_values: bool = True) -> str:
    """
    Formata uma descrição detalhada de um item de orçamento, incluindo atributos e componentes.

    Args:
        item: O objeto `ItemOrcamento` a ser formatado.
        include_monetary_values: Se `True`, inclui custos unitários dos componentes.

    Returns:
        Uma string contendo a descrição formatada do item.
    """
    display_name = ""
    componentes_str = ""

    if item.instancia:
        configuracao = item.instancia.configuracao
        display_name = configuracao.nome

        # Atributos da Instância
        numeric_attrs = []
        non_numeric_attrs = []
        for attr_instancia in item.instancia.atributos.all():
            if attr_instancia.template_atributo.atributo.tipo == 'num' and attr_instancia.valor_num is not None:
                numeric_attrs.append(str(int(attr_instancia.valor_num)))
            elif attr_instancia.template_atributo.atributo.tipo == 'str' and attr_instancia.valor_texto:
                non_numeric_attrs.append(attr_instancia.valor_texto)
        
        if non_numeric_attrs:
            display_name += f" - {' '.join(non_numeric_attrs)}"
        if numeric_attrs:
            display_name += f" ({'x'.join(numeric_attrs)})mm"

        # Componentes Calculados da Instância
        componentes_str = "\n" + _("--- Componentes ---") + "\n"
        for ic in item.instancia.componentes.all():
            component_line = f"- {ic.componente.nome}: {ic.quantidade} {ic.componente.unidade}"
            if include_monetary_values:
                component_line += f" (" + str(_("Custo Unit")) + f": {ic.custo_unitario})"
            component_line += "\n"
            if ic.descricao_detalhada:
                component_line += f"  " + str(_("Detalhes")) + f": {ic.descricao_detalhada}\n"
            componentes_str += component_line

    elif item.configuracao:
        display_name = item.configuracao.nome
        # Para itens que são apenas configurações (pais), podemos listar os componentes do template
        # mas sem quantidades calculadas, pois não há uma instância específica.
        componentes_str = "\n" + _("--- Componentes (Padrão) ---") + "\n"
        for tc in item.configuracao.template.componentes.all():
            # Tenta encontrar a escolha de componente real para esta configuração
            escolha = item.configuracao.componentes_escolha.filter(template_componente=tc).first()
            componente_nome = escolha.componente_real.nome if escolha else tc.componente.nome
            componentes_str += f"- {componente_nome}: {tc.quantidade_fixa or _('Variável')} {tc.componente.unidade}\n"
    else:
        display_name = _("Item de Orçamento Genérico")

    if item.codigo_item_manual:
        display_name = f"{item.codigo_item_manual} - {display_name}"

    return display_name + componentes_str


def _formatar_detalhes_item_orcamento(item: ItemOrcamento) -> str:
    """
    Formata os detalhes de um item de orçamento para exibição no Excel do orçamento.

    Args:
        item: O objeto `ItemOrcamento`.

    Returns:
        Uma string formatada com os detalhes do item e seus custos.
    """
    return _format_detailed_item_description_base(item, include_monetary_values=True)


def _formatar_detalhes_item_ficha_producao(item: ItemOrcamento) -> str:
    """
    Formata os detalhes de um item de orçamento para exibição na Ficha de Produção do Excel.

    Args:
        item: O objeto `ItemOrcamento`.

    Returns:
        Uma string formatada com os detalhes do item, sem incluir valores monetários.
    """
    return _format_detailed_item_description_base(item, include_monetary_values=False)


def _sanitize_name(name: str) -> str:
    """
    Sanitiza um nome para ser usado como variável de template, removendo acentos e caracteres especiais.

    Args:
        name: A string a ser sanitizada.

    Returns:
        A string sanitizada.
    """
    if not name:
        return ""
    # 1. Converter para minúsculas
    s = name.lower()
    # 2. Substituir caracteres acentuados comuns
    s = s.replace('á', 'a').replace('à', 'a').replace('â', 'a').replace('ã', 'a')
    s = s.replace('é', 'e').replace('ê', 'e')
    s = s.replace('í', 'i')
    s = s.replace('ó', 'o').replace('ô', 'o').replace('õ', 'o')
    s = s.replace('ú', 'u').replace('ü', 'u')
    s = s.replace('ç', 'c')
    # 3. Substituir sequências de não-alfanuméricos por '_'
    s = re.sub(r'[^a-z0-9]+', '_', s)
    # 4. Remover '_' no início/fim
    return s.strip('_')


def render_instancia_descricao(item_orcamento: ItemOrcamento) -> str:
    """
    Renderiza a descrição para uma linha de instância (nível 1.1.1) usando o template de instância.
    Foca-se nos atributos da instância.

    Args:
        item_orcamento: O objeto `ItemOrcamento` contendo a instância.

    Returns:
        Uma string com a descrição renderizada da instância.
    """
    if not item_orcamento.instancia:
        return _("Instância de item inválida")

    instancia = item_orcamento.instancia
    template_produto = instancia.configuracao.template
    template_str = template_produto.descricao_instancia_template

    # Fallback se não houver template: gera uma descrição simples dos atributos.
    if not template_str or "{{" not in template_str:
        numeric_attrs = []
        non_numeric_attrs = []
        for attr_instancia in instancia.atributos.all():
            if attr_instancia.template_atributo.atributo.tipo == 'num' and attr_instancia.valor_num is not None:
                numeric_attrs.append(str(int(attr_instancia.valor_num)))
            elif attr_instancia.template_atributo.atributo.tipo == 'str' and attr_instancia.valor_texto:
                non_numeric_attrs.append(attr_instancia.valor_texto)
        
        description = ' '.join(non_numeric_attrs)
        if numeric_attrs:
            description += f" ({'x'.join(numeric_attrs)})mm"
        return description.strip()

    # Construir contexto com atributos
    context_data = {}
    for ia in instancia.atributos.all():
        attr_name = _sanitize_name(ia.template_atributo.atributo.nome)
        valor = ia.valor_num if ia.template_atributo.atributo.tipo == 'num' else ia.valor_texto
        if isinstance(valor, decimal.Decimal) and valor == valor.to_integral_value():
            valor = int(valor)
        context_data[attr_name] = valor

    # Renderizar
    try:
        template = Template(template_str)
        context = Context(context_data)
        return template.render(context)
    except Exception as e:
        return _("[ERRO NO TEMPLATE DE INSTÂNCIA: {error}]").format(error=e)


def render_configuracao_descricao(configuracao: ProdutoConfiguracao) -> str:
    """
    Renderiza a descrição para uma linha de configuração (nível 1.1) usando o template de configuração.
    Foca-se nos componentes da configuração.

    Args:
        configuracao: O objeto `ProdutoConfiguracao`.

    Returns:
        Uma string com a descrição renderizada da configuração.
    """
    template_str = configuracao.descricao_configuracao_template

    # Fallback se não houver template: retorna o nome da configuração.
    if not template_str or "{{" not in template_str:
        return configuracao.nome

    # Construir contexto com componentes
    componentes_context = {}
    for escolha in configuracao.componentes_escolha.all():
        componente_template_name = _sanitize_name(escolha.template_componente.componente.nome)
        descricao_componente = escolha.descricao_personalizada or escolha.componente_real.nome
        componentes_context[componente_template_name] = descricao_componente
    
    context_data = {'componentes': componentes_context}

    # Renderizar
    try:
        template = Template(template_str)
        context = Context(context_data)
        return template.render(context)
    except Exception as e:
        return _("[ERRO NO TEMPLATE DE CONFIGURAÇÃO: {error}]").format(error=e)


def exportar_orcamento_excel(request: HttpRequest, orcamento_id: int, itens_orcamento: List[ItemOrcamento], total_geral_orcamento: float) -> HttpResponse:
    """
    Gera e serve um arquivo Excel para um orçamento específico.

    Este arquivo inclui uma estrutura hierárquica de itens, com categorias, templates
    e instâncias de produtos, além de cláusulas adicionais.

    Args:
        request: O objeto HttpRequest.
        orcamento_id: O ID do Orcamento a ser exportado.
        itens_orcamento: Uma lista de `ItemOrcamento` relacionados ao orçamento.
        total_geral_orcamento: O valor total geral do orçamento.

    Returns:
        Um HttpResponse contendo o arquivo .xlsx.

    Raises:
        FileNotFoundError: Se os arquivos de template Excel não forem encontrados.
        Exception: Para outros erros durante a geração do Excel.
    """
    orcamento = get_object_or_404(Orcamento, pk=orcamento_id)
    template_path = settings.BASE_DIR / 'static' / 'excel_templates' / 'modelo.xlsx'
    clauses_path = settings.BASE_DIR / 'static' / 'excel_templates' / 'modelo_clausulas.xlsx'

    try:
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active
        sheet['B3'] = orcamento.nome_cliente or ''
        sheet['B4'] = str(_("Obra")) + f": {orcamento.codigo_legado or ''}"
        sheet['B5'] = orcamento.codigo_legado or ''

        # Capturar estilos das linhas modelo
        category_model_row_styles = [copy(sheet.cell(row=9, column=col_idx)) for col_idx in range(1, 8)]
        template_model_row_styles = [copy(sheet.cell(row=10, column=col_idx)) for col_idx in range(1, 8)]
        instance_model_row_styles = [copy(sheet.cell(row=11, column=col_idx)) for col_idx in range(1, 8)]

        current_row = 9
        
        # --- Lógica de Agrupamento Hierárquico ---
        grouped_items = {}
        for item in itens_orcamento:
            if item.instancia and item.instancia.configuracao:
                config = item.instancia.configuracao
                categoria_nome = config.template.categoria.nome

                # Nível 1: Garantir que a Categoria existe no dicionário
                if categoria_nome not in grouped_items:
                    grouped_items[categoria_nome] = {}

                # Nível 2: Garantir que a Configuração existe na Categoria
                if config.id not in grouped_items[categoria_nome]:
                    grouped_items[categoria_nome][config.id] = {
                        'config_obj': config,
                        'instances': []
                    }
                
                # Adicionar a instância à lista correta
                grouped_items[categoria_nome][config.id]['instances'].append(item)

        # --- Lógica de Escrita no Excel ---
        category_counter = 0
        for categoria_nome, configs_data in grouped_items.items():
            category_counter += 1
            
            # Nível 1: Artigo (Categoria)
            sheet.insert_rows(current_row) # Insere uma nova linha na posição atual, empurrando o conteúdo existente para baixo
            for col_idx in range(1, 8):
                copy_style(category_model_row_styles[col_idx - 1], sheet.cell(row=current_row, column=col_idx))
            sheet.cell(row=current_row, column=1).value = f"{category_counter}"
            sheet.cell(row=current_row, column=2).value = categoria_nome
            current_row += 1

            config_counter = 0
            for config_id, config_data in configs_data.items():
                config_counter += 1
                config_obj = config_data['config_obj']
                instances = config_data['instances']

                # Nível 2: Template + Configuração
                sheet.insert_rows(current_row) # Insere uma nova linha na posição atual
                for col_idx in range(1, 8):
                    copy_style(template_model_row_styles[col_idx - 1], sheet.cell(row=current_row, column=col_idx))
                sheet.cell(row=current_row, column=1).value = f"{category_counter}.{config_counter}"
                sheet.cell(row=current_row, column=2).value = render_configuracao_descricao(config_obj)
                current_row += 1

                # Nível 3: Instância/Atributos
                instance_counter = 0
                for item in instances:
                    instance_counter += 1
                    sheet.insert_rows(current_row) # Insere uma nova linha na posição atual
                    for col_idx in range(1, 8):
                        copy_style(instance_model_row_styles[col_idx - 1], sheet.cell(row=current_row, column=col_idx))
                    
                    sheet.cell(row=current_row, column=1).value = f"{category_counter}.{config_counter}.{instance_counter}"
                    sheet.cell(row=current_row, column=2).value = render_instancia_descricao(item)
                    sheet.cell(row=current_row, column=3).value = item.instancia.configuracao.template.unidade or ''
                    sheet.cell(row=current_row, column=4).value = item.quantidade
                    sheet.cell(row=current_row, column=5).value = float(item.preco_unitario) if item.preco_unitario is not None else 0.0
                    sheet.cell(row=current_row, column=6).value = float(item.total) if item.total is not None else 0.0
                    current_row += 1

        # Deleta as linhas de modelo após a inserção do conteúdo dinâmico
        sheet.delete_rows(current_row, 3)

        # Carrega e anexa as cláusulas do arquivo modelo_clausulas.xlsx
        clauses_workbook = openpyxl.load_workbook(clauses_path)
        clauses_sheet = clauses_workbook.active
        
        row_offset = current_row - 1
        for r_idx, row in enumerate(clauses_sheet.iter_rows(), 1):
            for c_idx, source_cell in enumerate(row, 1):
                target_cell = sheet.cell(row=r_idx + row_offset, column=c_idx)
                copy_cell(source_cell, target_cell)

        for merged_range in clauses_sheet.merged_cells.ranges:
            min_col_letter = get_column_letter(merged_range.min_col)
            max_col_letter = get_column_letter(merged_range.max_col)
            new_range_string = f"{min_col_letter}{merged_range.min_row + row_offset}:{max_col_letter}{merged_range.max_row + row_offset}"
            sheet.merge_cells(new_range_string)

        sheet.cell(row=current_row, column=7).value = float(total_geral_orcamento) if total_geral_orcamento is not None else 0.0

        # Salva o workbook em um buffer de memória e retorna como HttpResponse
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="orcamento_{orcamento.codigo_legado}.xlsx"'
        return response

    except FileNotFoundError as e:
        messages.error(request, _("Ocorreu um erro: O arquivo {filename} não foi encontrado. Verifique se os templates 'modelo.xlsx' e 'modelo_clausulas.xlsx' estão no lugar certo.").format(filename=e.filename))
        return redirect('editar_orcamento', orcamento_id=orcamento_id)
    except Exception as e:
        messages.error(request, _("Erro ao exportar orçamento para Excel: {error}").format(error=e))
        return redirect('editar_orcamento', orcamento_id=orcamento_id)


def exportar_ficha_producao_excel(request: HttpRequest, orcamento: Orcamento, itens_orcamento: List[ItemOrcamento]) -> HttpResponse:
    """
    Gera e serve um arquivo Excel para a ficha de produção de um orçamento específico.

    Este arquivo inclui uma estrutura hierárquica de itens, com categorias, templates
    e instâncias de produtos, além de componentes agregados para a produção.

    Args:
        request: O objeto HttpRequest.
        orcamento: O objeto `Orcamento` a ser exportado.
        itens_orcamento: Uma lista de `ItemOrcamento` relacionados ao orçamento.

    Returns:
        Um HttpResponse contendo o arquivo .xlsx.

    Raises:
        FileNotFoundError: Se o arquivo de template Excel não for encontrado.
        Exception: Para outros erros durante a geração do Excel.
    """
    template_path = settings.BASE_DIR / 'static' / 'excel_templates' / 'modelo_ficha_producao.xlsx'

    try:
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active
        initial_max_row = sheet.max_row # Armazena o número máximo de linhas iniciais do template

        sheet['B3'] = orcamento.nome_cliente or ''
        sheet['B4'] = str(_("Obra")) + f": {orcamento.codigo_legado or ''}"
        sheet['B5'] = orcamento.codigo_legado or ''

        # Captura estilos das linhas modelo (assumindo que as linhas 9, 10, 11 são linhas modelo)
        category_model_row_styles = [copy(sheet.cell(row=9, column=col_idx)) for col_idx in range(1, 8)]
        aggregated_components_model_row_styles = [copy(sheet.cell(row=10, column=col_idx)) for col_idx in range(1, 8)]
        instance_model_row_styles = [copy(sheet.cell(row=11, column=col_idx)) for col_idx in range(1, 8)]

        # Deleta as linhas de modelo após capturar seus estilos
        # Deletando da linha de maior número primeiro para evitar problemas de deslocamento
        sheet.delete_rows(11, 1) # Deleta linha 11
        sheet.delete_rows(10, 1) # Deleta linha 10
        sheet.delete_rows(9, 1)  # Deleta linha 9

        # Ajusta current_row para refletir a nova posição inicial para o conteúdo
        # Se as linhas 9, 10, 11 foram deletadas, o conteúdo que começaria na 12 agora começa na 9
        current_row = 9
        
        # --- Lógica de Agrupamento Hierárquico ---
        grouped_items = {}
        for item in itens_orcamento:
            if item.instancia and item.instancia.configuracao:
                config = item.instancia.configuracao
                categoria_nome = config.template.categoria.nome

                if categoria_nome not in grouped_items:
                    grouped_items[categoria_nome] = {}

                if config.id not in grouped_items[categoria_nome]:
                    grouped_items[categoria_nome][config.id] = {
                        'config_obj': config,
                        'instances': [],
                        'aggregated_components': defaultdict(float)
                    }
                
                grouped_items[categoria_nome][config.id]['instances'].append(item)
                
                # Agregação de componentes para o Nível 1.1
                for ic in item.instancia.componentes.all():
                    # Usar uma tupla (nome, unidade, descricao_detalhada) como chave para agregar
                    component_key = (ic.componente.nome, ic.componente.unidade, ic.descricao_detalhada or '')
                    grouped_items[categoria_nome][config.id]['aggregated_components'][component_key] += float(ic.quantidade) * item.quantidade # Multiplica pela quantidade do item no orçamento

        # --- Lógica de Escrita no Excel ---
        category_counter = 0
        for categoria_nome, configs_data in grouped_items.items():
            category_counter += 1
            
            # Nível 1: Artigo (Categoria)
            sheet.insert_rows(current_row) # Insere uma nova linha na posição atual
            for col_idx in range(1, 8):
                copy_style(category_model_row_styles[col_idx - 1], sheet.cell(row=current_row, column=col_idx))
            sheet.cell(row=current_row, column=1).value = f"{category_counter}"
            sheet.cell(row=current_row, column=2).value = categoria_nome
            current_row += 1

            config_counter = 0
            for config_id, config_data in configs_data.items():
                config_counter += 1
                config_obj = config_data['config_obj']
                instances = config_data['instances']
                aggregated_components = config_data['aggregated_components']

                # Nível 1.1: Componentes Agregados
                sheet.insert_rows(current_row) # Insere uma nova linha na posição atual
                for col_idx in range(1, 8):
                    copy_style(aggregated_components_model_row_styles[col_idx - 1], sheet.cell(row=current_row, column=col_idx))
                sheet.cell(row=current_row, column=1).value = f"{category_counter}.{config_counter}"
                
                components_list_str = _("Componentes:") + "\n"
                for (comp_name, comp_unit, comp_desc), total_qty in aggregated_components.items():
                    unit_display = comp_unit
                    if comp_desc:
                        unit_display += f" - {comp_desc}"
                    line = f"- {comp_name}: {total_qty:.2f} {unit_display}"
                    components_list_str += line + "\n"
                
                cell = sheet.cell(row=current_row, column=2, value=components_list_str.strip())
                cell.alignment = Alignment(wrap_text=True)
                current_row += 1

                # Nível 1.1.1: Instância/Atributos
                instance_counter = 0
                for item in instances:
                    instance_counter += 1
                    sheet.insert_rows(current_row) # Insere uma nova linha na posição atual
                    for col_idx in range(1, 8):
                        copy_style(instance_model_row_styles[col_idx - 1], sheet.cell(row=current_row, column=col_idx))
                    
                    sheet.cell(row=current_row, column=1).value = f"{category_counter}.{config_counter}.{instance_counter}"
                    sheet.cell(row=current_row, column=2).value = render_instancia_descricao(item)
                    sheet.cell(row=current_row, column=3).value = item.instancia.configuracao.template.unidade or ''
                    sheet.cell(row=current_row, column=4).value = item.quantidade
                    current_row += 1
        
        # --- Limpeza e Adição de Conteúdo Final ---
        # Deleta todas as linhas desde current_row até o final original do template
        if current_row <= initial_max_row: # Apenas deleta se houver linhas para deletar
            sheet.delete_rows(current_row, initial_max_row - current_row + 1)

        # Carrega o modelo_final_ficha.xlsx
        final_ficha_path = settings.BASE_DIR / 'static' / 'excel_templates' / 'modelo_final_ficha.xlsx'
        final_ficha_workbook = openpyxl.load_workbook(final_ficha_path)
        final_ficha_sheet = final_ficha_workbook.active

        # Copia o conteúdo de modelo_final_ficha.xlsx (linhas 1-5, colunas A-G)
        row_offset_final_ficha = current_row - 1 # Ajusta o offset para inserção
        for r_idx in range(1, 6): # Linhas 1 a 5
            for c_idx in range(1, 8):
                source_cell = final_ficha_sheet.cell(row=r_idx, column=c_idx)
                target_cell = sheet.cell(row=r_idx + row_offset_final_ficha, column=c_idx)
                copy_cell(source_cell, target_cell)

        # Copia células mescladas de modelo_final_ficha.xlsx
        for merged_range in final_ficha_sheet.merged_cells.ranges:
            # Apenas copia se o range mesclado estiver dentro da área copiada (linhas 1-5, cols A-G)
            if merged_range.min_row >= 1 and merged_range.max_row <= 5 and \
               merged_range.min_col >= 1 and merged_range.max_col <= 7:
                min_col_letter = get_column_letter(merged_range.min_col)
                max_col_letter = get_column_letter(merged_range.max_col)
                new_range_string = f"{min_col_letter}{merged_range.min_row + row_offset_final_ficha}:{max_col_letter}{merged_range.max_row + row_offset_final_ficha}"
                sheet.merge_cells(new_range_string)

        # Aplica explicitamente a borda para a 5ª linha do conteúdo inserido
        # Isso corresponde ao sublinhado em modelo_final_ficha.xlsx
        underline_row_index = current_row + 4 # 5ª linha do conteúdo copiado
        thin_border = Border(bottom=Side(style='thin'))
        for col_idx in range(1, 8):
            cell = sheet.cell(row=underline_row_index, column=col_idx)
            cell.border = thin_border

        # Salva o workbook em um buffer de memória e retorna como HttpResponse
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ficha_producao_{orcamento.codigo_legado}.xlsx"'
        return response

    except FileNotFoundError:
        messages.error(request, _("O arquivo de template Excel para a ficha de produção (modelo_ficha_producao.xlsx) não foi encontrado."))
        return redirect('editar_orcamento', orcamento_id=orcamento.id)
    except Exception as e:
        messages.error(request, _("Erro ao exportar a ficha de produção: {error}").format(error=e))
        return redirect('editar_orcamento', orcamento_id=orcamento.id)